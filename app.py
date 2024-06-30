#DICIONARIO 
""" load_or_create_workbook(file_path):
Descripción: Carga un archivo Excel existente o crea uno nuevo si no existe.
Parámetros:
file_path (str): Ruta del archivo Excel.
Retorno:
Workbook: Libro de trabajo de Excel cargado o creado.
get_user_by_phone(phone_number, file_path):
Descripción: Busca un usuario por número de teléfono en el archivo Excel.
Parámetros:
phone_number (str): Número de teléfono a buscar.
file_path (str): Ruta del archivo Excel.
Retorno:
dict o None: Datos del usuario si se encuentra, None si no.
check_password_requirements(password):
Descripción: Verifica si la contraseña cumple con los requisitos necesarios.
Parámetros:
password (str): Contraseña a verificar.
Retorno:
bool: True si cumple con los requisitos, False en caso contrario.
generate_secure_password():
Descripción: Genera una contraseña segura que cumple con los requisitos.
Retorno:
str: Contraseña generada.
register_user(username, password, email, phone, file_path):
Descripción: Registra un nuevo usuario en el archivo Excel.
Parámetros:
username (str): Nombre de usuario.
password (str): Contraseña.
email (str): Correo electrónico.
phone (str): Número de teléfono.
file_path (str): Ruta del archivo Excel.
Retorno:
str: Mensaje de éxito o error.
is_phone_registered(phone_number, file_path):
Descripción: Verifica si un número de teléfono ya está registrado en el archivo Excel.
Parámetros:
phone_number (str): Número de teléfono a verificar.
file_path (str): Ruta del archivo Excel.
Retorno:
bool: True si el número de teléfono está registrado, False si no.
get_user_by_username(username, file_path):
Descripción: Busca un usuario por nombre de usuario en el archivo Excel.
Parámetros:
username (str): Nombre de usuario a buscar.
file_path (str): Ruta del archivo Excel.
Retorno:
dict o None: Datos del usuario si se encuentra, None si no.
Rutas de la Aplicación:
/verify-phone: Verifica si un número de teléfono está registrado.
/generate-password: Genera una contraseña segura.
/forgot-account: Página para recuperar cuentas olvidadas.
/register: Registra un nuevo usuario.
/login: Inicia sesión con nombre de usuario y contraseña.
/: Página de inicio. """
# Flask es un framework web ligero para Python. Se utiliza para crear aplicaciones web y APIs.
from flask import Flask, request, jsonify, render_template, redirect, url_for
# El módulo 're' se utiliza para trabajar con patrones de texto.
# Puedes usarlo para buscar, extraer y manipular cadenas de caracteres basadas en patrones específicos.
import re
# El módulo 'secrets' se utiliza para generar números aleatorios seguros, como tokens de autenticación o contraseñas temporales.
import secrets
# El módulo 'string' proporciona una colección de constantes con caracteres ASCII, como letras mayúsculas, minúsculas, dígitos y caracteres de puntuación.
# Puede ser útil para manipular cadenas de caracteres.
import string
# openpyxl es una biblioteca para trabajar con archivos de Excel (xlsx).
# Puedes leer, escribir y modificar hojas de cálculo de Excel utilizando esta biblioteca.
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
# El módulo 'os' proporciona funciones para interactuar con el sistema operativo.
# Puedes usarlo para acceder a variables de entorno, crear directorios, manipular rutas de archivos y más.
import os

app = Flask(__name__)


def load_or_create_workbook(file_path):
    """
    Carga un archivo Excel existente o crea uno nuevo si no existe.

    Args:
        file_path (str): Ruta del archivo Excel.

    Returns:
        Workbook: Libro de trabajo de Excel cargado o creado.
    """
    if os.path.exists(file_path):
        return load_workbook(file_path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre de Usuario', 'Contraseña', 'Correo Electrónico', 'Número de Teléfono'])
        wb.save(file_path)
        return wb

# Definir los encabezados de las columnas
headers = ['Nombre de Usuario', 'Contraseña', 'Correo Electrónico', 'Número de Teléfono']

# Cargar el archivo Excel existente o crear uno nuevo si no existe
file_path = 'cuentas.xlsx'
wb = load_or_create_workbook(file_path)
ws = wb.active
def get_user_by_phone(phone_number, file_path):
    """
    Busca un usuario por número de teléfono en el archivo Excel.

    Args:
        phone_number (str): Número de teléfono a buscar.
        file_path (str): Ruta del archivo Excel.

    Returns:
        dict or None: Un diccionario con los datos del usuario si se encuentra, o None si no se encuentra.
            - 'username': Nombre de usuario
            - 'password': Contraseña
            - 'email': Correo electrónico
            - 'phone': Número de teléfono
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if phone_number == row[3]:  # Asumiendo que el teléfono está en la cuarta columna
            return {'username': row[0], 'password': row[1], 'email': row[2], 'phone': row[3]}
    return None

# Expresión regular para validar direcciones de correo electrónico
EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@(gmail\.com|outlook\.com)$')

# Requisitos para contraseñas
password_requirements = {
    'uppercase': 1,
    'digits': 1,
    'special': 1,
    'min_length': 8
}

def check_password_requirements(password):
    """
    Verifica si la contraseña cumple con los requisitos necesarios.

    Args:
        password (str): Contraseña a verificar.

    Returns:
        bool: True si cumple con los requisitos, False en caso contrario.
    """
    patterns = {
        'uppercase': r'[A-Z]',
        'digits': r'\d',
        'special': r'\W'
    }
    return all(len(re.findall(pattern, password)) >= password_requirements[requirement]
               for requirement, pattern in patterns.items()) and len(password) >= password_requirements['min_length']

def generate_secure_password():
    """
    Genera una contraseña segura que cumple con los requisitos.

    Returns:
        str: Contraseña generada.
    """
    alphabet = string.ascii_letters + string.digits + string.punctuation
    while True:
        password = ''.join(secrets.choice(alphabet) for i in range(password_requirements['min_length']))
        if check_password_requirements(password):
            break
    return password

# Función para registrar un nuevo usuario
def register_user(username, password, email, phone, file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    # Verificar si el nombre de usuario ya existe
    for row in ws.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
        if username in row:
            return 'El nombre de usuario ya está en uso. Por favor, elige otro.'
    # Verificar si el correo electrónico ya está en uso
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True):
        if email in row:
            return 'Correo Electrónico ya en uso. Por favor, utiliza otro correo.'
    # Agregar el nuevo usuario
    ws.append([username, password, email, phone])
    wb.save(file_path)
    return 'Usuario registrado con éxito.'

def is_phone_registered(phone_number, file_path):
    """
    Verifica si un número de teléfono ya está registrado en el archivo Excel.

    Args:
        phone_number (str): Número de teléfono a verificar.
        file_path (str): Ruta del archivo Excel.

    Returns:
        bool: True si el número de teléfono está registrado, False si no.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if phone_number == row[3]:  # Asumiendo que el teléfono está en la cuarta columna
            return True
    return False


def get_user_by_username(username, file_path):
    """
    Busca un usuario por nombre de usuario en el archivo Excel.

    Args:
        username (str): Nombre de usuario a buscar.
        file_path (str): Ruta del archivo Excel.

    Returns:
        dict or None: Un diccionario con los datos del usuario si se encuentra, o None si no se encuentra.
            - 'username': Nombre de usuario
            - 'password': Contraseña
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if username == row[0]:  # Asumiendo que el nombre de usuario está en la primera columna
            return {'username': row[0], 'password': row[1]}
    return None

@app.route('/verify-phone', methods=['POST'])
def verify_phone():
    phone_number = request.form['phone'].strip()
    user = get_user_by_phone(phone_number, 'cuentas.xlsx')
    if user:
        # Aquí iría la lógica para enviar el código de verificación al número
        # Por ejemplo, usando un servicio de SMS
        return render_template('enter_code.html')  # Página para ingresar el código
    else:
        return jsonify({'error': 'Número de teléfono no registrado.'}), 400
@app.route('/generate-password', methods=['POST'])
def generate_password():
    return jsonify({'password': generate_secure_password()})

@app.route('/forgot-account')
def forgot_account():
    return render_template('forgotaccount.html')
@app.route('/register', methods=['POST']) 
def register(): 
    name = request.form['name'].strip() 
    password = request.form['password'] 
    phone = request.form['phone'].strip() 
    email = request.form['email'].strip() 
    file_path = 'cuentas.xlsx' 
    wb = load_workbook(file_path) 
    ws = wb.active 
    
    # Verificar si el nombre de usuario ya está en uso
    for row in ws.iter_rows(min_row=2, values_only=True): 
        if name == row[0]: 
            return jsonify({'error': 'El nombre de usuario ya está en uso. Por favor, elige otro.'}), 400 
    
    # Verificar si el correo electrónico ya está en uso
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, values_only=True): 
        if email == row[0]: 
            return jsonify({'error': 'Correo Electrónico ya en uso. Por favor, utiliza otro correo.'}), 400 
    
    # Verificar si el número de teléfono ya está en uso
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=2, values_only=True): 
        if phone == row[0]: 
            return jsonify({'error': 'Número de teléfono ya en uso. Por favor, utiliza otro número.'}), 400 
    
    if not check_password_requirements(password): 
        return jsonify({'error': 'La contraseña no cumple con los requisitos necesarios.'}), 400 
    if not name or not password or not phone or not email: 
        return jsonify({'error': 'Todos los campos son obligatorios.'}), 400 
    if not EMAIL_REGEX.match(email): 
        return jsonify({'error': 'Por favor, ingresa un correo electrónico válido de Gmail o Outlook.'}), 400 
    
    ws.append([name, password, email, phone]) 
    wb.save(file_path) 
    return jsonify({'message': 'Usuario registrado con éxito.'})
@app.route('/login', methods=['POST'])
def login():
    username = request.form['username'].strip()
    password = request.form['password']
    file_path = 'cuentas.xlsx'
    user = get_user_by_username(username, file_path)
    if not user:
        return jsonify({'error': 'El nombre de usuario no está registrado. Por favor, verifica tus datos o regístrate.'}), 400
    if user['password'] != password:
        return jsonify({'error': 'La contraseña introducida es incorrecta. Por favor, inténtalo de nuevo.'}), 400
    return render_template('login.html')

@app.route('/login_page')
def login_page():
    # Cargar el archivo Excel
    file_path = 'cuentas.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active

    # Obtener el nombre de usuario del primer usuario en el archivo Excel (solo como ejemplo)
    username = ws.cell(row=2, column=1).value  # Asumiendo que el nombre de usuario está en la primera columna

    return render_template('login.html', username=username)


@app.route('/')
def index():
    return render_template('index.html')
@app.route('/send_verification', methods=['POST'])
def send_verification():
    # Aquí procesas la solicitud y envías el código al correo
    # Luego, rediriges al usuario a la página principal
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
